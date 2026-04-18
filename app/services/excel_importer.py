"""Service: import Avito Excel exports into products + product_images.

Avito's Excel export contains one sheet per category ("Мужская обувь-Кроссовки"
etc.) plus reference sheets ("Спр-*") and an "Инструкция" sheet (skipped).

Per row: match by avito_id first, then by case-insensitive trimmed title (for
products with avito_id IS NULL). UPDATE brand/goods_type/goods_subtype/size/
color/price/image_url; replace product_images with the URLs from
"Ссылки на фото" (split by " | ").
"""

from __future__ import annotations

import os
from datetime import datetime, timedelta, timezone
from io import BytesIO

import openpyxl
import structlog
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings

from app.models.account import Account
from app.models.product import Product
from app.models.product_image import ProductImage

logger = structlog.get_logger(__name__)


def _delete_local_image_files(image_urls: list[str]) -> int:
    """Delete local files for image URLs starting with /media/.

    Returns the number of files successfully deleted.
    Skips external URLs and logs warnings for missing files.
    """
    deleted = 0
    for url in image_urls:
        if not url.startswith("/media/"):
            continue
        # /media/products/123/file.jpg → <MEDIA_DIR>/products/123/file.jpg
        rel = url[len("/media/"):]
        filepath = os.path.normpath(os.path.join(settings.MEDIA_DIR, rel))
        media_root = os.path.normpath(settings.MEDIA_DIR)
        if not filepath.startswith(media_root):
            continue
        try:
            os.remove(filepath)
            deleted += 1
        except FileNotFoundError:
            logger.warning("excel_importer.orphan_file_missing", path=filepath)
        except OSError as e:
            logger.warning("excel_importer.orphan_delete_failed", path=filepath, error=str(e))
    return deleted


SHEET_SKIP_PREFIXES = ("Спр-",)
SHEET_SKIP_NAMES = {"Инструкция"}

# Avito column header → Product attribute. Resilient to column reordering.
COLUMN_MAP = {
    "Номер объявления на Авито": "avito_id",
    "Название объявления": "title",
    "Описание объявления": "description",
    "Цена": "price",
    "Бренд одежды": "brand",
    "Вид одежды": "goods_type",
    "Вид одежды, обуви, аксессуаров": "goods_subtype",
    "Размер": "size",
    "Цвет": "color",
    "Категория": "category",
    "Состояние": "condition",
    "Ссылки на фото": "_photos",          # special: needs splitting
    "AvitoDateEnd": "_avito_date_end",    # special: ISO datetime, used to derive published_at
    "AvitoStatus": "_avito_status",       # special: mirrored into product.extra for reference
}


class InvalidExcelError(Exception):
    """Raised when the file isn't a valid xlsx Avito export."""


# ── pure helpers ──────────────────────────────────────────────────────────────

def _norm(s: str | None) -> str:
    return " ".join((s or "").split()).lower()


def _str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _int(val) -> int | None:
    if val is None or val == "":
        return None
    try:
        return int(float(str(val).strip()))
    except (TypeError, ValueError):
        return None


def _normalize_avito_image_url(url: str) -> str:
    """Convert Avito-internal autoload-feed image URLs into public CDN URLs.

    Avito Excel exports list photo URLs as
        http(s)://avito.ru/autoload/N/items-to-feed/images?imageSlug=/image/...
    These are rate-limited and 301-redirect to the autoload admin area,
    not viewable in a browser. The actual public CDN URL is
        https://NN.img.avito.st/image/...
    where NN is any 2-digit shard (00-99 all work). We pick "00".

    Plain HTTP URLs are upgraded to HTTPS to avoid mixed-content blocking
    on our HTTPS pages.
    """
    if not url:
        return url
    # Translate the autoload-feed wrapper URL to a CDN URL
    if "/autoload/" in url and "imageSlug=" in url:
        slug = url.split("imageSlug=", 1)[1]
        if slug.startswith("/"):
            return "https://00.img.avito.st" + slug
    # Otherwise just upgrade scheme
    if url.startswith("http://"):
        return "https://" + url[len("http://"):]
    return url


def _split_photos(raw: str | None) -> list[str]:
    """Split "|"-separated photo URLs, filter to http(s), normalize to public CDN."""
    if not raw:
        return []
    parts = [p.strip() for p in raw.split("|")]
    return [_normalize_avito_image_url(p) for p in parts if p.startswith("http")]


def _parse_avito_date_end(raw) -> datetime | None:
    """Parse Avito's AvitoDateEnd value into a naive UTC datetime, then subtract
    30 days to derive the publication date.

    Avito ads run for 30 days from publication, so:
        published_at = AvitoDateEnd - 30 days

    Input formats:
      - ISO 8601 string with offset, e.g. "2026-05-11T20:43:59+03:00"
      - datetime instance (openpyxl may return one if the cell is date-typed)

    Returns a tz-naive UTC datetime, or None if missing/unparseable.
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        dt = raw
    else:
        s = str(raw).strip()
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s)
        except (TypeError, ValueError):
            return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt - timedelta(days=30)


def _parse_workbook_bytes(file_bytes: bytes) -> list[dict]:
    """Parse xlsx bytes → list of row-dicts. Skips Spravochnik / instruction sheets.

    Layout: row 1 = category title, row 2 = headers, rows 3-4 = metadata,
    rows 5+ = data.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise InvalidExcelError(f"Не удалось открыть xlsx: {e}") from e

    rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in SHEET_SKIP_NAMES:
            continue
        if any(sheet_name.startswith(p) for p in SHEET_SKIP_PREFIXES):
            continue
        ws = wb[sheet_name]
        if ws.max_row < 5:
            continue

        headers = [c.value for c in ws[2]]

        for ri in range(5, ws.max_row + 1):
            row = ws[ri]
            row_vals = [c.value for c in row]
            if not any(v not in (None, "") for v in row_vals):
                continue
            entry: dict = {"_sheet": sheet_name}
            for header, val in zip(headers, row_vals):
                if header is None:
                    continue
                entry[str(header).strip()] = val
            if not entry.get("Номер объявления на Авито") and not entry.get("Название объявления"):
                continue
            rows.append(entry)
    return rows


def _row_to_updates(row: dict) -> dict:
    """Map an Excel row → product field updates (filtered to non-empty values)."""
    out: dict = {}
    for header, attr in COLUMN_MAP.items():
        if attr.startswith("_"):
            continue
        raw = row.get(header)
        if raw is None or raw == "":
            continue
        if attr in ("price", "avito_id"):
            v = _int(raw)
            if v is not None:
                out[attr] = v
        else:
            v = _str(raw)
            if v is not None:
                out[attr] = v[:255]

    # description has no length cap
    if "description" in out and row.get("Описание объявления"):
        out["description"] = str(row["Описание объявления"]).strip()

    photos = _split_photos(row.get("Ссылки на фото"))
    if photos:
        out["_photos"] = photos
        out["image_url"] = photos[0][:500]

    # AvitoDateEnd → derived published_at (applied conditionally in import_avito_excel:
    # only set when the product currently has no published_at).
    pub_at = _parse_avito_date_end(row.get("AvitoDateEnd"))
    if pub_at is not None:
        out["_published_at"] = pub_at

    # AvitoStatus → product.extra["avito_status_excel"] (always refreshed)
    avito_status = _str(row.get("AvitoStatus"))
    if avito_status:
        out["_avito_status_excel"] = avito_status[:255]

    return out


# ── main entry point ──────────────────────────────────────────────────────────

async def import_avito_excel(
    account_id: int,
    file_bytes: bytes,
    db: AsyncSession,
) -> dict:
    """Read xlsx bytes, sync products + product_images for this account.

    Returns counters: {updated, created, photos, skipped, errors}.
    Raises InvalidExcelError if the file can't be parsed at all.
    """
    counters = {"updated": 0, "created": 0, "photos": 0, "skipped": 0, "errors": 0}

    account = await db.get(Account, account_id)
    if not account:
        counters["errors"] = 1
        return counters

    rows = _parse_workbook_bytes(file_bytes)
    if not rows:
        return counters

    # Snapshot existing products for this account (matching indexes)
    existing_result = await db.execute(
        select(Product).where(Product.account_id == account_id)
    )
    all_products = list(existing_result.scalars().all())
    by_avito = {p.avito_id: p for p in all_products if p.avito_id is not None}
    by_title_null_avito: dict[str, Product] = {}
    for p in all_products:
        if p.avito_id is None and p.title:
            by_title_null_avito.setdefault(_norm(p.title), p)

    # Global avito_ids (avoid unique-constraint conflicts across accounts)
    global_result = await db.execute(
        select(Product.avito_id).where(Product.avito_id.isnot(None))
    )
    global_avito_ids = {row[0] for row in global_result.all()}

    now = datetime.now(timezone.utc).replace(tzinfo=None)

    for row in rows:
        try:
            upd = _row_to_updates(row)
            avito_id = upd.get("avito_id")
            title = upd.get("title")

            # ── Match ──
            target: Product | None = None
            if avito_id and avito_id in by_avito:
                target = by_avito[avito_id]
            elif title:
                target = by_title_null_avito.get(_norm(title))

            if target is None:
                if not avito_id and not title:
                    counters["skipped"] += 1
                    continue
                if avito_id and avito_id in global_avito_ids:
                    logger.warning(
                        "excel_importer.foreign_avito_id",
                        account_id=account_id, avito_id=avito_id,
                    )
                    counters["skipped"] += 1
                    continue
                target = Product(
                    account_id=account_id,
                    title=(title or f"[Авито] {avito_id}")[:255],
                    status="imported",
                    # Prefer the AvitoDateEnd-derived value when available;
                    # fall back to "now" so the row still has a sortable timestamp.
                    published_at=upd.get("_published_at") or now,
                )
                db.add(target)
                await db.flush()  # need target.id for product_images
                if avito_id:
                    global_avito_ids.add(avito_id)
                    by_avito[avito_id] = target
                counters["created"] += 1
            else:
                counters["updated"] += 1

            # ── Apply scalar updates ──
            photos = upd.pop("_photos", None)
            published_at_val = upd.pop("_published_at", None)
            avito_status_excel = upd.pop("_avito_status_excel", None)
            for k, v in upd.items():
                # Don't overwrite an existing avito_id with a different one
                if k == "avito_id" and target.avito_id and target.avito_id != v:
                    continue
                setattr(target, k, v)

            # When AvitoDateEnd is present in the row, treat it as authoritative
            # and overwrite any existing value (which is usually a placeholder
            # 'now' set at first import). When the column is missing/empty we
            # leave the current value alone.
            if published_at_val is not None:
                target.published_at = published_at_val

            # Mirror AvitoStatus into JSONB extra for reference / debugging.
            if avito_status_excel is not None:
                extra = dict(target.extra or {})
                extra["avito_status_excel"] = avito_status_excel
                target.extra = extra

            # ── Replace photos ──
            if photos:
                old_imgs = await db.execute(
                    select(ProductImage.url).where(ProductImage.product_id == target.id)
                )
                old_urls = [r[0] for r in old_imgs.all()]
                _delete_local_image_files(old_urls)
                await db.execute(
                    delete(ProductImage).where(ProductImage.product_id == target.id)
                )
                for sort_order, url in enumerate(photos[:10]):
                    db.add(ProductImage(
                        product_id=target.id,
                        url=url[:500],
                        filename=f"avito_{sort_order}",
                        sort_order=sort_order,
                        is_main=(sort_order == 0),
                    ))
                counters["photos"] += min(len(photos), 10)

            await db.commit()
        except Exception as e:
            await db.rollback()
            counters["errors"] += 1
            logger.warning(
                "excel_importer.row_error",
                account_id=account_id, error=str(e),
            )

    logger.info(
        "excel_importer.done",
        account_id=account_id, account=account.name, **counters, total_rows=len(rows),
    )
    return counters
