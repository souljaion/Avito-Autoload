"""Sync products from fresh Avito cabinet Excel export.

Reads an .xlsx export from Avito cabinet, matches rows to existing
products by avito_id, and creates/updates/deletes to bring DB in sync
with what's actually live on Avito.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime, timezone

import openpyxl
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.product import Product
from app.models.product_image import ProductImage
from app.models.listing import Listing

logger = logging.getLogger(__name__)

COLUMN_MAP = {
    "Уникальный идентификатор объявления": "feed_ad_id",
    "Номер объявления на Авито": "avito_id",
    "Название объявления": "title",
    "Описание объявления": "description",
    "Цена": "price",
    "Бренд одежды": "brand",
    "Вид одежды": "goods_type",
    "Вид одежды, обуви, аксессуаров": "subcategory",
    "Тип товара": "subcategory",
    "Размер": "size",
    "Цвет": "color",
    "Категория": "category",
    "Состояние": "condition",
    "Ссылки на фото": "_photos",
    "AvitoStatus": "_avito_status",
}


@dataclass
class SyncReport:
    to_create: int = 0
    to_update: int = 0
    to_delete: int = 0
    excel_rows_total: int = 0
    excel_rows_skipped: int = 0
    examples_create: list[dict] = field(default_factory=list)
    examples_update: list[dict] = field(default_factory=list)
    examples_delete: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _parse_excel(excel_path: str) -> tuple[list[dict], list[str]]:
    """Parse Avito cabinet Excel export, return (rows, errors)."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    rows: list[dict] = []
    errors: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Инструкция" or sheet_name.startswith("Спр-"):
            continue

        ws = wb[sheet_name]
        ws_rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(ws_rows) < 5:
            continue

        header_row = ws_rows[1]  # row 2 (0-indexed: 1)
        if not header_row:
            continue

        headers = [str(h).strip() if h else "" for h in header_row]

        col_indices: dict[str, int] = {}
        for col_name in COLUMN_MAP:
            for i, h in enumerate(headers):
                if h == col_name:
                    col_indices[col_name] = i
                    break

        if "Номер объявления на Авито" not in col_indices:
            errors.append(f"Sheet '{sheet_name}': column 'Номер объявления на Авито' not found")
            continue

        for row_idx, row_data in enumerate(ws_rows[4:], start=5):  # data from row 5
            if not row_data or all(c is None for c in row_data):
                continue

            row_dict: dict[str, str | None] = {}
            for col_name, attr_name in COLUMN_MAP.items():
                idx = col_indices.get(col_name)
                if idx is not None and idx < len(row_data):
                    val = row_data[idx]
                    row_dict[attr_name] = str(val).strip() if val is not None else None
                else:
                    row_dict[attr_name] = None

            row_dict["_sheet"] = sheet_name
            row_dict["_row"] = str(row_idx)
            rows.append(row_dict)

    wb.close()
    return rows, errors


async def sync_from_excel(
    excel_path: str,
    account_id: int,
    db: AsyncSession,
    dry_run: bool = True,
) -> SyncReport:
    """Sync products for account_id from Avito cabinet Excel export."""
    report = SyncReport()

    rows, parse_errors = _parse_excel(excel_path)
    report.errors.extend(parse_errors)

    # Filter active rows with avito_id
    active_rows: list[dict] = []
    for row in rows:
        report.excel_rows_total += 1

        avito_id_raw = row.get("avito_id")
        if not avito_id_raw:
            report.excel_rows_skipped += 1
            continue

        try:
            avito_id_int = int(avito_id_raw)
        except (ValueError, TypeError):
            report.errors.append(
                f"Sheet {row.get('_sheet')} row {row.get('_row')}: invalid avito_id '{avito_id_raw}'"
            )
            report.excel_rows_skipped += 1
            continue

        status = row.get("_avito_status", "")
        if status and status != "Активно":
            report.excel_rows_skipped += 1
            continue

        row["_avito_id_int"] = avito_id_int
        active_rows.append(row)

    # Load existing products for this account
    result = await db.execute(
        select(Product).where(Product.account_id == account_id)
    )
    existing_products = {p.avito_id: p for p in result.scalars().all() if p.avito_id}

    avito_ids_in_excel = set()
    products_to_create: list[dict] = []
    products_to_update: list[tuple[Product, dict]] = []

    for row in active_rows:
        avito_id = row["_avito_id_int"]
        avito_ids_in_excel.add(avito_id)

        feed_ad_id = row.get("feed_ad_id") or ""
        title = row.get("title") or f"Product {avito_id}"
        description = row.get("description")
        price_raw = row.get("price")
        price = None
        if price_raw:
            try:
                price = int(float(price_raw))
            except (ValueError, TypeError):
                pass

        update_data = {
            "avito_id": avito_id,
            "feed_ad_id": feed_ad_id if feed_ad_id else str(avito_id),
            "title": title,
            "description": description,
            "price": price,
            "brand": row.get("brand"),
            "goods_type": row.get("goods_type"),
            "category": row.get("category"),
            "subcategory": row.get("subcategory"),
            "size": row.get("size"),
            "color": row.get("color"),
            "condition": row.get("condition"),
        }

        existing = existing_products.get(avito_id)
        if existing:
            products_to_update.append((existing, update_data))
        else:
            products_to_create.append(update_data)

    # Determine orphans (in DB but not in Excel)
    orphan_products: list[Product] = []
    for avito_id, product in existing_products.items():
        if avito_id not in avito_ids_in_excel:
            orphan_products.append(product)

    report.to_create = len(products_to_create)
    report.to_update = len(products_to_update)
    report.to_delete = len(orphan_products)

    report.examples_create = [
        {"avito_id": d["avito_id"], "title": d["title"], "price": d["price"]}
        for d in products_to_create[:5]
    ]
    report.examples_update = [
        {"id": p.id, "avito_id": p.avito_id, "title": d["title"], "feed_ad_id": d["feed_ad_id"]}
        for p, d in products_to_update[:5]
    ]
    report.examples_delete = [
        {"id": p.id, "avito_id": p.avito_id, "title": p.title}
        for p in orphan_products[:10]
    ]

    if dry_run:
        return report

    # Apply changes
    now = datetime.now(timezone.utc).replace(tzinfo=None)

    for data in products_to_create:
        product = Product(
            avito_id=data["avito_id"],
            feed_ad_id=data["feed_ad_id"],
            title=data["title"],
            description=data["description"],
            price=data["price"],
            brand=data["brand"],
            goods_type=data["goods_type"],
            category=data["category"],
            subcategory=data["subcategory"],
            size=data["size"],
            color=data["color"],
            condition=data["condition"],
            status="imported",
            account_id=account_id,
            published_at=now,
        )
        db.add(product)

    for product, data in products_to_update:
        product.feed_ad_id = data["feed_ad_id"]
        product.title = data["title"]
        product.description = data["description"]
        product.price = data["price"]
        product.brand = data["brand"]
        product.goods_type = data["goods_type"]
        product.category = data["category"]
        product.subcategory = data["subcategory"]
        product.size = data["size"]
        product.color = data["color"]
        product.condition = data["condition"]
        product.status = "imported"
        product.removed_at = None
        product.updated_at = now

    if orphan_products:
        orphan_ids = [p.id for p in orphan_products]
        await db.execute(delete(Product).where(Product.id.in_(orphan_ids)))

    await db.flush()

    logger.info(
        "sync_from_excel applied: account_id=%d created=%d updated=%d deleted=%d",
        account_id, report.to_create, report.to_update, report.to_delete,
    )

    return report
