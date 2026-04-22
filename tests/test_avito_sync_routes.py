"""Tests for /admin/avito-sync routes."""

import io
import os

import openpyxl
import pytest
import pytest_asyncio
from unittest.mock import AsyncMock, patch, MagicMock

from app.services.sync_from_avito_export import SyncReport


def _make_xlsx_bytes(rows: list[dict] | None = None) -> bytes:
    """Create minimal Avito export xlsx in memory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Обувь"
    headers = [
        "Уникальный идентификатор объявления",
        "Номер объявления на Авито",
        "Название объявления",
        "Описание объявления",
        "Цена",
        "AvitoStatus",
    ]
    ws.cell(1, 1, "Обувь")
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    ws.cell(3, 1, "Обязательный")
    ws.cell(4, 1, "Подробнее")
    if rows:
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, h in enumerate(headers, 1):
                if h in row_data:
                    ws.cell(row_idx, col_idx, row_data[h])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestAvitoSyncRoutes:

    @pytest.mark.asyncio
    async def test_get_page_returns_200(self, client):
        resp = await client.get("/admin/avito-sync")
        assert resp.status_code == 200
        assert "Синхронизация с Авито" in resp.text

    @pytest.mark.asyncio
    async def test_preview_with_valid_file(self, client):
        xlsx_bytes = _make_xlsx_bytes([
            {
                "Уникальный идентификатор объявления": "123",
                "Номер объявления на Авито": "999888777",
                "Название объявления": "Test",
                "Цена": "5000",
                "AvitoStatus": "Активно",
            },
        ])
        resp = await client.post(
            "/admin/avito-sync/preview",
            data={"account_id": "1"},
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert "Предпросмотр" in resp.text

    @pytest.mark.asyncio
    async def test_apply_without_preview_returns_400(self, client):
        xlsx_bytes = _make_xlsx_bytes()
        resp = await client.post(
            "/admin/avito-sync/apply",
            data={"account_id": "1", "preview_hash": ""},
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_apply_with_wrong_hash_returns_400(self, client):
        xlsx_bytes = _make_xlsx_bytes()
        resp = await client.post(
            "/admin/avito-sync/apply",
            data={"account_id": "1", "preview_hash": "wronghash"},
            files={"file": ("test.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 400
