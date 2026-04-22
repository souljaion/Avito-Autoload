"""Tests for sync_from_avito_export service."""

import os
import uuid

import openpyxl
import pytest
import pytest_asyncio
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine

from app.config import settings
from app.models.product import Product
from app.models.product_image import ProductImage
from app.models.listing import Listing
from app.services.sync_from_avito_export import sync_from_excel


ACCOUNT_ID = 1


def _make_excel(path: str, rows: list[dict], *, sheet_name: str = "Обувь") -> str:
    """Create a minimal Avito cabinet export Excel file."""
    wb = openpyxl.Workbook()

    # Instruction sheet (should be skipped)
    ws_instr = wb.active
    ws_instr.title = "Инструкция"
    ws_instr.cell(1, 1, "Пропусти меня")

    # Data sheet
    ws = wb.create_sheet(sheet_name)
    headers = [
        "Уникальный идентификатор объявления",
        "Номер объявления на Авито",
        "Название объявления",
        "Описание объявления",
        "Цена",
        "Бренд одежды",
        "Вид одежды",
        "Вид одежды, обуви, аксессуаров",
        "Размер",
        "Цвет",
        "Категория",
        "Состояние",
        "Ссылки на фото",
        "AvitoStatus",
    ]
    # Row 1: category name
    ws.cell(1, 1, "Обувь")
    # Row 2: headers
    for i, h in enumerate(headers, 1):
        ws.cell(2, i, h)
    # Row 3: required/optional
    ws.cell(3, 1, "Обязательный")
    # Row 4: hints
    ws.cell(4, 1, "Подробнее")
    # Rows 5+: data
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, h in enumerate(headers, 1):
            key = h
            if key in row_data:
                ws.cell(row_idx, col_idx, row_data[key])

    filepath = os.path.join(path, "test_export.xlsx")
    wb.save(filepath)
    wb.close()
    return filepath


@pytest_asyncio.fixture
async def sync_db():
    """Isolated DB session that commits and rolls back via savepoint."""
    eng = create_async_engine(str(settings.DATABASE_URL))
    async with eng.connect() as conn:
        trans = await conn.begin()
        session = AsyncSession(bind=conn, expire_on_commit=False)
        yield session
        await session.close()
        if trans.is_active:
            await trans.rollback()
    await eng.dispose()


class TestSyncFromAvitoExport:

    @pytest.mark.asyncio
    async def test_dry_run_no_changes(self, sync_db, tmp_path):
        """dry_run=True should not modify DB."""
        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": "123",
                "Номер объявления на Авито": "999000111",
                "Название объявления": "Test Sneaker",
                "Цена": "5000",
                "AvitoStatus": "Активно",
            },
        ])

        # Count products before
        result = await sync_db.execute(
            select(Product).where(Product.account_id == ACCOUNT_ID)
        )
        before_count = len(result.scalars().all())

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=True)

        assert report.to_create >= 0
        # Verify no actual changes
        result = await sync_db.execute(
            select(Product).where(Product.account_id == ACCOUNT_ID)
        )
        after_count = len(result.scalars().all())
        assert before_count == after_count

    @pytest.mark.asyncio
    async def test_create_new_products(self, sync_db, tmp_path):
        """Import 2 new products into empty account."""
        # Use a unique avito_id to avoid conflicts
        aid1 = 770000001
        aid2 = 770000002
        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": "feed_id_1",
                "Номер объявления на Авито": str(aid1),
                "Название объявления": "Nike Air Max",
                "Описание объявления": "Cool shoes",
                "Цена": "5000",
                "Бренд одежды": "Nike",
                "AvitoStatus": "Активно",
            },
            {
                "Уникальный идентификатор объявления": "feed_id_2",
                "Номер объявления на Авито": str(aid2),
                "Название объявления": "Adidas Ultraboost",
                "Цена": "7000",
                "AvitoStatus": "Активно",
            },
        ])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=False)

        assert report.to_create == 2
        assert report.to_update == 0

        result = await sync_db.execute(
            select(Product).where(Product.avito_id == aid1)
        )
        p = result.scalar_one()
        assert p.title == "Nike Air Max"
        assert p.feed_ad_id == "feed_id_1"
        assert p.status == "imported"
        assert p.brand == "Nike"
        assert p.price == 5000

    @pytest.mark.asyncio
    async def test_update_existing_by_avito_id(self, sync_db, tmp_path):
        """Existing product matched by avito_id gets feed_ad_id updated."""
        avito_id = 770000003
        product = Product(
            avito_id=avito_id,
            title="Old Title",
            price=3000,
            status="imported",
            account_id=ACCOUNT_ID,
        )
        sync_db.add(product)
        await sync_db.flush()
        product_id = product.id

        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": "avito_feed_999",
                "Номер объявления на Авито": str(avito_id),
                "Название объявления": "New Title",
                "Цена": "4000",
                "Бренд одежды": "Puma",
                "AvitoStatus": "Активно",
            },
        ])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=False)

        assert report.to_update == 1
        assert report.to_create == 0

        result = await sync_db.execute(
            select(Product).where(Product.id == product_id)
        )
        p = result.scalar_one()
        assert p.feed_ad_id == "avito_feed_999"
        assert p.title == "New Title"
        assert p.price == 4000
        assert p.brand == "Puma"

    @pytest.mark.asyncio
    async def test_delete_orphans_not_in_excel(self, sync_db, tmp_path):
        """Products in DB but not in Excel should be deleted."""
        keep_id = 770000010
        del_id1 = 770000011
        del_id2 = 770000012

        for aid in [keep_id, del_id1, del_id2]:
            sync_db.add(Product(
                avito_id=aid, title=f"P{aid}", price=1000,
                status="imported", account_id=ACCOUNT_ID,
            ))
        await sync_db.flush()

        # Excel only has keep_id
        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": str(keep_id),
                "Номер объявления на Авито": str(keep_id),
                "Название объявления": "Keeper",
                "Цена": "1000",
                "AvitoStatus": "Активно",
            },
        ])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=False)

        assert report.to_delete == 2
        assert report.to_update == 1

        result = await sync_db.execute(
            select(Product).where(Product.avito_id.in_([del_id1, del_id2]))
        )
        assert len(result.scalars().all()) == 0

        result = await sync_db.execute(
            select(Product).where(Product.avito_id == keep_id)
        )
        assert result.scalar_one() is not None

    @pytest.mark.asyncio
    async def test_skip_non_active_status(self, sync_db, tmp_path):
        """Rows with AvitoStatus != 'Активно' are skipped."""
        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": "111",
                "Номер объявления на Авито": "770000020",
                "Название объявления": "Active",
                "Цена": "1000",
                "AvitoStatus": "Активно",
            },
            {
                "Уникальный идентификатор объявления": "222",
                "Номер объявления на Авито": "770000021",
                "Название объявления": "Archived",
                "Цена": "2000",
                "AvitoStatus": "Архив",
            },
        ])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=True)

        assert report.to_create == 1
        assert report.excel_rows_skipped == 1

    @pytest.mark.asyncio
    async def test_skip_row_without_avito_id(self, sync_db, tmp_path):
        """Rows missing avito_id are skipped."""
        excel_path = _make_excel(str(tmp_path), [
            {
                "Уникальный идентификатор объявления": "333",
                "Название объявления": "No ID",
                "Цена": "1000",
                "AvitoStatus": "Активно",
            },
        ])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=True)

        assert report.to_create == 0
        assert report.excel_rows_skipped == 1

    @pytest.mark.asyncio
    async def test_cascade_delete_removes_product_images(self, sync_db, tmp_path):
        """When product is deleted, its product_images are cascade-deleted by DB."""
        avito_id = 770000030
        product = Product(
            avito_id=avito_id, title="Will be deleted", price=1000,
            status="imported", account_id=ACCOUNT_ID,
        )
        sync_db.add(product)
        await sync_db.flush()

        img1 = ProductImage(
            product_id=product.id, url="/media/test1.jpg",
            filename="test1.jpg", sort_order=0, is_main=True,
        )
        img2 = ProductImage(
            product_id=product.id, url="/media/test2.jpg",
            filename="test2.jpg", sort_order=1, is_main=False,
        )
        sync_db.add_all([img1, img2])
        await sync_db.flush()
        pid = product.id

        # Excel is empty (no rows) → product should be deleted
        excel_path = _make_excel(str(tmp_path), [])

        report = await sync_from_excel(excel_path, ACCOUNT_ID, sync_db, dry_run=False)

        assert report.to_delete >= 1

        result = await sync_db.execute(
            select(ProductImage).where(ProductImage.product_id == pid)
        )
        assert len(result.scalars().all()) == 0
