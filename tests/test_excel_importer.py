"""Tests for app/services/excel_importer.py + POST /accounts/{id}/import-excel."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest
from fastapi import FastAPI
from httpx import AsyncClient, ASGITransport

from app.db import get_db
from app.models.product import Product
from app.models.product_image import ProductImage
from app.routes.accounts import router as accounts_router
from app.services.excel_importer import (
    InvalidExcelError,
    _int,
    _norm,
    _parse_workbook_bytes,
    _row_to_updates,
    _split_photos,
    _str,
    import_avito_excel,
)


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------

class TestNorm:
    def test_lowercases_and_collapses(self):
        assert _norm("  Hello   WORLD  ") == "hello world"

    def test_none(self):
        assert _norm(None) == ""

    def test_empty(self):
        assert _norm("") == ""


class TestStr:
    def test_strips(self):
        assert _str("  abc  ") == "abc"

    def test_none(self):
        assert _str(None) is None

    def test_empty_after_strip(self):
        assert _str("    ") is None

    def test_int_stringified(self):
        assert _str(42) == "42"


class TestInt:
    @pytest.mark.parametrize("val,expected", [
        (None, None),
        ("", None),
        ("42", 42),
        ("4990.0", 4990),
        (4990, 4990),
        (4990.7, 4990),
        (" 100 ", 100),
        ("not-a-number", None),
        ("abc123", None),
    ])
    def test_parse(self, val, expected):
        assert _int(val) == expected


class TestSplitPhotos:
    def test_pipe_separated_upgrades_http_to_https(self):
        # http:// is upgraded to https:// to avoid mixed-content blocking on
        # our HTTPS pages. https:// stays as-is.
        raw = "http://a/1 | http://b/2 | https://c/3"
        assert _split_photos(raw) == ["https://a/1", "https://b/2", "https://c/3"]

    def test_filter_non_http(self):
        raw = "http://a | not-a-url | https://b"
        assert _split_photos(raw) == ["https://a", "https://b"]

    def test_empty(self):
        assert _split_photos(None) == []
        assert _split_photos("") == []

    def test_trailing_pipe(self):
        # Avito exports often have trailing " | " with empty last segment
        assert _split_photos("http://a | http://b | ") == ["https://a", "https://b"]

    def test_translates_avito_autoload_url_to_cdn(self):
        # Avito Excel uses an internal autoload-feed wrapper URL that 429's in
        # the browser. We translate it to a public CDN URL.
        raw = ("https://avito.ru/autoload/1/items-to-feed/images"
               "?imageSlug=/image/1/1.HASH | "
               "http://avito.ru/autoload/1/items-to-feed/images"
               "?imageSlug=/image/1/2.OTHER")
        assert _split_photos(raw) == [
            "https://00.img.avito.st/image/1/1.HASH",
            "https://00.img.avito.st/image/1/2.OTHER",
        ]


# ---------------------------------------------------------------------------
# Fixture: build minimal Avito-shaped workbook in-memory
# ---------------------------------------------------------------------------

HEADERS = [
    "Уникальный идентификатор объявления", "Способ размещения",
    "Номер объявления на Авито", "Номер телефона", "Адрес",
    "Ссылки на фото", "Способ связи", "Название объявления",
    "Описание объявления", "Категория", "Цена", "Вид одежды",
    "Состояние", "Вид объявления", "Бренд одежды", "Цвет",
    "Цвет от производителя", "Материал основной части",
    "Соединять", "Название мульти", "Вид одежды, обуви, аксессуаров",
    "Размер", "Целевая аудитория", "AvitoDateEnd", "AvitoStatus",
]
REQUIRED_ROW = ["Обязательный"] * len(HEADERS)
HINTS_ROW = ["Подробнее"] * len(HEADERS)


def _data_row(avito_id, title, brand="Nike", price="5000",
              photos="http://avito.ru/img1.jpg | http://avito.ru/img2.jpg",
              goods_type="Мужская обувь", goods_subtype="Кроссовки",
              size="42", color="Белый"):
    row = [None] * len(HEADERS)
    row[0] = str(avito_id)
    row[2] = str(avito_id)
    row[5] = photos
    row[7] = title
    row[10] = price
    row[11] = goods_type
    row[14] = brand
    row[15] = color
    row[20] = goods_subtype
    row[21] = size
    return row


def _build_xlsx_bytes(sheets: dict[str, list[list]]) -> bytes:
    """Build an xlsx and return its raw bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for ri, row in enumerate(rows, 1):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _full_sheet(name: str, *data_rows) -> tuple[str, list[list]]:
    return name, [
        ["Личные вещи > Одежда"],  # row 1
        HEADERS,                    # row 2
        REQUIRED_ROW,               # row 3
        HINTS_ROW,                  # row 4
        *data_rows,                 # row 5+
    ]


# ---------------------------------------------------------------------------
# _parse_workbook_bytes
# ---------------------------------------------------------------------------

class TestParseWorkbookBytes:
    def test_skips_instructions_and_reference_sheets(self):
        xb = _build_xlsx_bytes(dict([
            ("Инструкция", [["Some text"]]),
            ("Спр-Мужская обувь-Кеды", [["Reference"]]),
            _full_sheet("Мужская обувь-Кроссовки", _data_row(101, "Nike Air Max")),
        ]))
        rows = _parse_workbook_bytes(xb)
        assert len(rows) == 1
        assert rows[0]["Номер объявления на Авито"] == "101"
        assert rows[0]["_sheet"] == "Мужская обувь-Кроссовки"

    def test_reads_multiple_data_rows(self):
        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки",
                _data_row(1, "First"),
                _data_row(2, "Second"),
                _data_row(3, "Third"),
            ),
        ]))
        rows = _parse_workbook_bytes(xb)
        assert [r["Номер объявления на Авито"] for r in rows] == ["1", "2", "3"]

    def test_skips_empty_rows(self):
        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки",
                _data_row(1, "First"),
                [None] * len(HEADERS),
                _data_row(2, "Second"),
            ),
        ]))
        rows = _parse_workbook_bytes(xb)
        assert len(rows) == 2

    def test_skips_rows_without_id_and_title(self):
        bad = [None] * len(HEADERS)
        bad[10] = "5000"  # only price
        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", bad),
        ]))
        assert _parse_workbook_bytes(xb) == []

    def test_combines_multiple_data_sheets(self):
        xb = _build_xlsx_bytes(dict([
            _full_sheet("Мужская обувь-Кроссовки", _data_row(1, "Sneaker")),
            _full_sheet("Мужская одежда-Кофты", _data_row(2, "Hoodie")),
        ]))
        rows = _parse_workbook_bytes(xb)
        sheets = {r["_sheet"] for r in rows}
        assert sheets == {"Мужская обувь-Кроссовки", "Мужская одежда-Кофты"}

    def test_invalid_xlsx_raises(self):
        with pytest.raises(InvalidExcelError):
            _parse_workbook_bytes(b"definitely not an xlsx file")

    def test_empty_bytes_raises(self):
        with pytest.raises(InvalidExcelError):
            _parse_workbook_bytes(b"")


# ---------------------------------------------------------------------------
# _row_to_updates
# ---------------------------------------------------------------------------

class TestRowToUpdates:
    def test_full_row(self):
        row = {
            "Номер объявления на Авито": "7989201866",
            "Название объявления": "Nike Air Max 90",
            "Описание объявления": "Кроссовки оригинал",
            "Цена": "5990",
            "Бренд одежды": "Nike",
            "Вид одежды": "Мужская обувь",
            "Вид одежды, обуви, аксессуаров": "Кроссовки",
            "Размер": "42",
            "Цвет": "Белый",
            "Ссылки на фото": "http://a/1.jpg | http://b/2.jpg",
        }
        out = _row_to_updates(row)
        assert out["avito_id"] == 7989201866
        assert out["title"] == "Nike Air Max 90"
        assert out["price"] == 5990
        assert out["brand"] == "Nike"
        assert out["goods_type"] == "Мужская обувь"
        assert out["goods_subtype"] == "Кроссовки"
        assert out["size"] == "42"
        assert out["color"] == "Белый"
        # Generic non-avito URLs still get https-upgrade (no autoload wrapper)
        assert out["image_url"] == "https://a/1.jpg"
        assert out["_photos"] == ["https://a/1.jpg", "https://b/2.jpg"]
        assert out["description"] == "Кроссовки оригинал"

    def test_empty_fields_omitted(self):
        out = _row_to_updates({
            "Номер объявления на Авито": "111",
            "Название объявления": "X",
            "Бренд одежды": None,
            "Вид одежды": "",
        })
        assert out == {"avito_id": 111, "title": "X"}

    def test_invalid_price_omitted(self):
        out = _row_to_updates({
            "Номер объявления на Авито": "111",
            "Цена": "договорная",
        })
        assert "price" not in out
        assert out["avito_id"] == 111

    def test_truncates_to_255(self):
        out = _row_to_updates({
            "Номер объявления на Авито": "1",
            "Бренд одежды": "X" * 500,
        })
        assert len(out["brand"]) == 255

    def test_no_photos_no_image_url(self):
        out = _row_to_updates({
            "Номер объявления на Авито": "1",
            "Ссылки на фото": "",
        })
        assert "image_url" not in out
        assert "_photos" not in out


# ---------------------------------------------------------------------------
# import_avito_excel — service entry point
# ---------------------------------------------------------------------------

def _make_account(id=3, name="Zulla"):
    a = MagicMock()
    a.id = id
    a.name = name
    return a


def _make_product(id, account_id=3, avito_id=None, title=None, brand=None):
    p = MagicMock()
    p.id = id
    p.account_id = account_id
    p.avito_id = avito_id
    p.title = title
    p.brand = brand
    return p


def _build_db(account, existing=None, global_avito_ids=None):
    """Mock async session with the 2 SELECTs the importer makes up front,
    then any number of subsequent execute() calls (DELETEs)."""
    existing = existing or []
    global_ids = global_avito_ids or []

    db = AsyncMock()
    db.get = AsyncMock(return_value=account)
    db.commit = AsyncMock()
    db.rollback = AsyncMock()
    db.flush = AsyncMock()

    added: list = []
    next_id = [10000]

    def _add(obj):
        added.append(obj)
        if not getattr(obj, "id", None):
            obj.id = next_id[0]
            next_id[0] += 1
    db.add = MagicMock(side_effect=_add)

    existing_scalars = MagicMock()
    existing_scalars.all.return_value = existing
    existing_result = MagicMock()
    existing_result.scalars.return_value = existing_scalars

    global_result = MagicMock()
    global_result.all.return_value = [(x,) for x in global_ids]

    delete_result = MagicMock()
    seq = [existing_result, global_result]

    async def _exec(*args, **kwargs):
        return seq.pop(0) if seq else delete_result
    db.execute = AsyncMock(side_effect=_exec)

    db.added = added
    return db


class TestImportAvitoExcel:
    @pytest.mark.asyncio
    async def test_account_not_found(self):
        db = AsyncMock()
        db.get = AsyncMock(return_value=None)
        # Real bytes don't matter — function returns before parsing
        counters = await import_avito_excel(999, b"any", db)
        assert counters == {"updated": 0, "created": 0, "photos": 0, "skipped": 0, "errors": 1}

    @pytest.mark.asyncio
    async def test_match_by_avito_id_updates(self):
        account = _make_account(id=3)
        existing = _make_product(id=42, avito_id=111, title="Old", brand=None)
        db = _build_db(account, existing=[existing], global_avito_ids=[111])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(111, "New title", brand="Nike")),
        ]))
        counters = await import_avito_excel(3, xb, db)

        assert counters["updated"] == 1
        assert counters["created"] == 0
        assert counters["errors"] == 0
        assert existing.title == "New title"
        assert existing.brand == "Nike"
        assert existing.goods_type == "Мужская обувь"
        # No new product
        new_products = [o for o in db.added if isinstance(o, Product)]
        assert new_products == []

    @pytest.mark.asyncio
    async def test_match_by_title_when_avito_id_null(self):
        account = _make_account(id=3)
        existing = _make_product(id=99, avito_id=None, title="Nike Air Max")
        db = _build_db(account, existing=[existing], global_avito_ids=[])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(222, "  NIKE  AIR  MAX  ", brand="Nike")),
        ]))
        counters = await import_avito_excel(3, xb, db)

        assert counters["updated"] == 1
        assert existing.avito_id == 222

    @pytest.mark.asyncio
    async def test_creates_new_when_no_match(self):
        account = _make_account(id=3)
        db = _build_db(account, existing=[], global_avito_ids=[])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(333, "Brand new", brand="Adidas")),
        ]))
        counters = await import_avito_excel(3, xb, db)

        assert counters["created"] == 1
        assert counters["updated"] == 0
        new_products = [o for o in db.added if isinstance(o, Product)]
        assert len(new_products) == 1
        new_p = new_products[0]
        assert new_p.avito_id == 333
        assert new_p.title == "Brand new"
        assert new_p.brand == "Adidas"
        assert new_p.status == "imported"
        assert new_p.account_id == 3

    @pytest.mark.asyncio
    async def test_skips_avito_id_owned_by_other_account(self):
        account = _make_account(id=3)
        # 444 is held globally by another account
        db = _build_db(account, existing=[], global_avito_ids=[444])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(444, "Stolen?", brand="Nike")),
        ]))
        counters = await import_avito_excel(3, xb, db)

        assert counters["skipped"] == 1
        assert counters["created"] == 0
        new_products = [o for o in db.added if isinstance(o, Product)]
        assert new_products == []

    @pytest.mark.asyncio
    async def test_photos_create_product_images(self):
        account = _make_account(id=3)
        existing = _make_product(id=42, avito_id=555, title="Has photos")
        db = _build_db(account, existing=[existing], global_avito_ids=[555])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(555, "Has photos")),
        ]))
        counters = await import_avito_excel(3, xb, db)

        assert counters["photos"] == 2  # _data_row default has 2 URLs
        images = [o for o in db.added if isinstance(o, ProductImage)]
        assert len(images) == 2
        assert images[0].is_main is True
        assert images[0].sort_order == 0
        assert images[1].is_main is False
        assert images[1].sort_order == 1
        assert existing.image_url.startswith("http")

    @pytest.mark.asyncio
    async def test_photos_dont_duplicate_on_second_import(self):
        """Second import for same product replaces images via DELETE then INSERT."""
        account = _make_account(id=3)
        existing = _make_product(id=42, avito_id=555, title="Has photos")
        db = _build_db(account, existing=[existing], global_avito_ids=[555])

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(555, "Has photos")),
        ]))
        await import_avito_excel(3, xb, db)
        # Verify a DELETE statement was sent before the inserts
        # The _build_db mock returns delete_result for any execute() after the first 2
        # We should see at least one execute() call beyond the initial 2 (the DELETE)
        assert db.execute.call_count >= 3  # 2 selects + 1 delete

    @pytest.mark.asyncio
    async def test_invalid_xlsx_raises(self):
        account = _make_account(id=3)
        db = _build_db(account)
        with pytest.raises(InvalidExcelError):
            await import_avito_excel(3, b"not an xlsx file", db)


# ---------------------------------------------------------------------------
# POST /accounts/{id}/import-excel — endpoint
# ---------------------------------------------------------------------------

def _make_app(mock_db):
    app = FastAPI()
    app.include_router(accounts_router)

    async def _gen():
        yield mock_db
    app.dependency_overrides[get_db] = _gen
    return app


class TestEndpoint:
    @pytest.mark.asyncio
    async def test_returns_200_with_counters(self):
        account = _make_account(id=3)
        db = _build_db(account, existing=[], global_avito_ids=[])
        app = _make_app(db)

        xb = _build_xlsx_bytes(dict([
            _full_sheet("Кроссовки", _data_row(700, "Brand new")),
        ]))

        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
            resp = await c.post(
                "/accounts/3/import-excel",
                files={"file": ("zulla.xlsx", xb,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()
        assert data["ok"] is True
        assert data["created"] == 1
        assert data["updated"] == 0
        assert data["photos"] == 2
        assert data["errors"] == 0

    @pytest.mark.asyncio
    async def test_404_for_unknown_account(self):
        db = AsyncMock()
        db.get = AsyncMock(return_value=None)
        app = _make_app(db)

        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
            resp = await c.post(
                "/accounts/999/import-excel",
                files={"file": ("x.xlsx", b"x", "application/octet-stream")},
            )
        assert resp.status_code == 404
        assert resp.json()["ok"] is False

    @pytest.mark.asyncio
    async def test_400_for_non_xlsx(self):
        account = _make_account(id=3)
        db = _build_db(account)
        app = _make_app(db)

        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
            resp = await c.post(
                "/accounts/3/import-excel",
                files={"file": ("data.csv", b"a,b,c", "text/csv")},
            )
        assert resp.status_code == 400
        assert "xlsx" in resp.json()["error"].lower()

    @pytest.mark.asyncio
    async def test_400_for_empty_file(self):
        account = _make_account(id=3)
        db = _build_db(account)
        app = _make_app(db)

        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
            resp = await c.post(
                "/accounts/3/import-excel",
                files={"file": ("zulla.xlsx", b"", "application/octet-stream")},
            )
        assert resp.status_code == 400

    @pytest.mark.asyncio
    async def test_400_for_invalid_xlsx_content(self):
        """A .xlsx-named file with garbage content → InvalidExcelError → 400."""
        account = _make_account(id=3)
        db = _build_db(account)
        app = _make_app(db)

        async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
            resp = await c.post(
                "/accounts/3/import-excel",
                files={"file": ("fake.xlsx", b"not really xlsx", "application/octet-stream")},
            )
        assert resp.status_code == 400
        assert "xlsx" in resp.json()["error"].lower() or "не удалось" in resp.json()["error"].lower()
